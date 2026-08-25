# -*- coding: utf-8 -*-
"""
查询服务层 — 桌面版和 Web 版共用的业务逻辑
从 widgets/query_tab.py 中提取，去除 PyQt5 依赖
"""
import os
import datetime
import logging

from form_parser import FormParser
from core.param_service import (
    build_sql_with_params, normalize_params, resolve_default, static_options
)

logger = logging.getLogger('DBQuery.query_service')


# ════════════════════════════════════════
#  SQL 构建
# ════════════════════════════════════════

def escape_sql_param(value):
    """转义参数值中的单引号，防止 SQL 语法错误"""
    return value.replace("'", "''")


def build_final_sql(form, param_values, options_by_name=None, already_normalized=False):
    """根据表单定义构建最终 SQL。

    只接受 ``form.params`` 中声明的参数名；未知客户端字段不会参与替换。动态
    select 的候选项由调用方通过 ``options_by_name`` 提供，避免将 label 或任意
    搜索词误当成 SQL 参数。
    """
    normalized = param_values if already_normalized else normalize_params(
        form, param_values, options_by_name=options_by_name
    )
    return build_sql_with_params(form, normalized)


# ════════════════════════════════════════
#  结果数据过滤/排序（纯数据，无 UI 依赖）
# ════════════════════════════════════════

class ResultData:
    """查询结果的纯数据表示，支持过滤和排序"""

    def __init__(self, columns, rows):
        self.columns = columns
        self.rows = rows

    def filter(self, text, col_idx=-1):
        """
        过滤结果
        :param text: 过滤关键字（大小写不敏感）
        :param col_idx: 列索引，-1 表示全部列
        :return: 新的 ResultData
        """
        if not text:
            return ResultData(self.columns, list(self.rows))

        text_lower = text.lower()
        filtered = []
        for row in self.rows:
            if col_idx >= 0:
                if col_idx < len(row):
                    cell = str(row[col_idx] or '').lower()
                    if text_lower in cell:
                        filtered.append(row)
            else:
                for val in row:
                    cell = str(val or '').lower()
                    if text_lower in cell:
                        filtered.append(row)
                        break

        return ResultData(self.columns, filtered)

    def sort(self, col_idx, descending=False):
        """
        排序结果
        :param col_idx: 列索引
        :param descending: 是否降序
        :return: 新的 ResultData
        """
        def sort_key(row):
            v = row[col_idx] if col_idx < len(row) else None
            if v is None:
                return (1, '')
            return (0, str(v))

        sorted_rows = sorted(self.rows, key=sort_key, reverse=descending)
        return ResultData(self.columns, sorted_rows)

    @property
    def row_count(self):
        return len(self.rows)

    @property
    def col_count(self):
        return len(self.columns)


# ════════════════════════════════════════
#  Excel 导出（纯 openpyxl，无 PyQt5）
# ════════════════════════════════════════

def export_to_excel(path, columns, rows, form_title='', form_desc='',
                    elapsed=0.0, params_info=None, final_sql=''):
    """
    将查询结果导出为 Excel 文件
    :param path: 保存路径
    :param columns: 列名列表
    :param rows: 数据行列表
    :param form_title: 表单标题
    :param form_desc: 表单描述
    :param elapsed: 查询耗时（秒）
    :param params_info: [(label, value), ...] 查询参数信息
    :param final_sql: 最终执行的 SQL
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if params_info is None:
        params_info = []

    wb = openpyxl.Workbook()

    # ── Sheet1: 查询结果 ──
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = u"查询结果"

    hdr_fill  = PatternFill("solid", fgColor="1A6EB5")
    hdr_font  = Font(color="FFFFFF", bold=True, name="微软雅黑", size=10)
    even_fill = PatternFill("solid", fgColor="EEF4FC")
    thin      = Side(style='thin', color='BBBBBB')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal='center', vertical='center')
    vcenter   = Alignment(vertical='center')

    # 表头
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.border    = border
        cell.alignment = center

    # 数据行
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = vcenter
            if ri % 2 == 0:
                cell.fill = even_fill

    # 自适应列宽（采样前 100 行）
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(len(rows) + 2, 102)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = 'A2'

    # 自动筛选
    try:
        last_col_letter = get_column_letter(len(columns))
        last_row = len(rows) + 1
        ws.auto_filter.ref = "A1:{}{}".format(last_col_letter, last_row)
    except Exception:
        pass

    # ── Sheet2: 查询信息 ──
    ws2 = wb.create_sheet(u"查询信息")
    info_rows = [
        (u"查询项目",   form_title),
        (u"项目说明",   form_desc or ''),
        (u"查询时间",   datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        (u"导出记录数", len(rows)),
        (u"字段数",     len(columns)),
        (u"查询耗时",   u"{:.2f} 秒".format(elapsed)),
        ('', ''),
        (u"—— 查询条件 ——", ''),
    ]
    info_rows += params_info
    # info_rows += [
    #     ('', ''),
    #     (u"—— SQL 语句 ——", ''),
    #     (final_sql, ''),
    # ]
    for r_idx, (k, v) in enumerate(info_rows, 1):
        ws2.cell(row=r_idx, column=1, value=k)
        if v != '':
            ws2.cell(row=r_idx, column=2, value=v)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 50

    wb.save(path)
    logger.info("Excel exported to: %s", path)


# ════════════════════════════════════════
#  参数值序列化（供 Web API 使用）
# ════════════════════════════════════════

def serialize_form(form, base_dir=None):
    """将 QueryForm 序列化为 dict（JSON 友好）"""
    # 将绝对路径转为相对路径（URL 安全，用正斜杠）
    fp = form.file_path
    if base_dir:
        try:
            fp = os.path.relpath(fp, base_dir)
        except ValueError:
            pass
    fp = fp.replace('\\', '/')

    return {
        'title':       form.title,
        'description': form.description,
        'group':       form.group,
        'query_type':  form.query_type,
        'web_enabled': bool(getattr(form, 'web_enabled', False)),
        'file_path':   fp,

        'params': [
            {
                'name':    p.name,
                'label':   p.label,
                'ptype':   p.ptype,
                # 保留旧 options 字段，同时提供 value/label 分离的 option_items。
                'options': p.options,
                'option_items': static_options(p),
                'default': resolve_default(p),
                'raw_default': p.default,
                'placeholder': p.placeholder,
                'required': p.required,
                'width': p.width,
                # 浏览器仅知道是否需要请求服务端加载动态候选；SQL 永不下发。
                'dynamic_options': bool(p.options_sql),
                'searchable': p.searchable or p.ptype == 'select',
                'allow_custom': p.allow_custom,

            }
            for p in form.params
        ],
    }


def load_all_forms(forms_dir, web_only=False):
    """加载所有表单，返回按分组组织的 dict（JSON 友好）。

    ``web_only=True`` 时只返回 [meta] 中明确 ``web_enabled = true`` 的表单。
    默认值保持桌面端兼容：桌面仍可加载和管理全部本地表单。
    """
    raw = FormParser.load_forms_from_dir(forms_dir)
    base_dir = os.path.dirname(forms_dir)  # 上级目录，路径相对于它（包含 forms/）
    result = {}
    for group, forms in raw.items():
        visible_forms = [
            form for form in forms
            if not web_only or bool(getattr(form, 'web_enabled', False))
        ]
        if visible_forms:
            result[group] = [serialize_form(f, base_dir) for f in visible_forms]
    return result
